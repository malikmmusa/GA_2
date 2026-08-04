#!/usr/bin/env python3
"""
Generate an HTML report comparing the app's predicted fovea-to-GA distance
against Gold Standard manual ImageJ measurements.

Reads:
  - Gold Standard Measurements from ImageJ.xlsx  (manual µm distances; bold = include)
  - test_validation/summary.csv                  (pred_distance_px per image)
  - test_validation/bland_altman_data.csv         (gt_disc_height_px for px→µm conversion)

Outputs:
  - test_validation/gold_standard_report.html

Matching logic:
  - Bold "Initial distance" → summary row with pair_mode = paired_before or single
  - Bold "Final distance"   → summary row with pair_mode = paired_after (_2 image)
  - Conversion: app_distance_um = pred_distance_px * 1800 / gt_disc_height_px
"""

from __future__ import annotations

import argparse
import base64
import csv
import statistics
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

DISC_DIAMETER_MICRONS = 1800.0
GOLD_STANDARD_XLSX = Path("/Users/musamalik/Desktop/Gold Standard Measurements from ImageJ.xlsx")

# Manually verified app distances (µm) that override the pipeline computation.
# Used when the bland_altman disc height is from a different coordinate space
# than the image loaded in the app, causing incorrect px→µm scaling.
APP_UM_OVERRIDES: dict[str, float] = {
    "22973051": 290.8,  # app shows 290.8 µm (disc 187.5 px); bland_altman has disc 360 px (wrong scale)
}


def load_csv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def encode_image(path: Path) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("ascii")


def safe_float(v) -> float | None:
    try:
        return float(v) if v is not None and str(v).strip() not in ("", "nan") else None
    except (ValueError, AttributeError):
        return None


def err_class_um(um: float | None) -> str:
    if um is None:
        return ""
    if um <= 50:
        return "row-good"
    if um <= 150:
        return "row-warn"
    return "row-bad"


def err_color_um(um: float | None) -> str:
    if um is None:
        return ""
    if um <= 50:
        return "err-good"
    if um <= 150:
        return "err-warn"
    return "err-bad"


def fmt_um(v: float | None) -> str:
    return f"{v:.0f}" if v is not None else "—"


def fmt_pct(v: float | None) -> str:
    return f"{v:.1f}%" if v is not None else "—"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=PROJECT_ROOT / "test_validation")
    parser.add_argument(
        "--gold-standard",
        type=Path,
        default=GOLD_STANDARD_XLSX,
    )
    args = parser.parse_args()

    out_dir = args.output_dir
    gold_path = args.gold_standard
    summary_path = out_dir / "summary.csv"
    ba_path = out_dir / "bland_altman_data.csv"

    if not gold_path.exists():
        print(f"ERROR: Gold Standard file not found: {gold_path}")
        return 1
    if not summary_path.exists():
        print(f"ERROR: {summary_path} not found.")
        return 1

    # ── Load summary (keyed by image_filename stem) ──────────────────────────
    summary_rows = load_csv(summary_path)
    # Build two lookups: case_key → rows for before/single, and for after
    before_lookup: dict[str, dict] = {}   # case_key → paired_before or single row
    after_lookup: dict[str, dict] = {}    # case_key → paired_after row
    for row in summary_rows:
        if row.get("prediction_detected", "no") != "yes":
            continue
        ck = str(row["case_key"])
        mode = row.get("pair_mode", "")
        if mode in ("paired_before", "single"):
            before_lookup[ck] = row
        elif mode == "paired_after":
            after_lookup[ck] = row

    # ── Load disc heights from bland_altman_data.csv ──────────────────────────
    disc_height_lookup: dict[str, float] = {}
    if ba_path.exists():
        for row in load_csv(ba_path):
            iid = row.get("image_id", "").strip()
            dh = safe_float(row.get("gt_disc_height_px"))
            if iid and dh:
                disc_height_lookup[iid] = dh

    def get_disc_height(image_filename: str, case_key: str) -> float | None:
        stem = Path(image_filename).stem
        return disc_height_lookup.get(stem) or disc_height_lookup.get(case_key)

    # ── Parse Gold Standard Excel (bold detection via openpyxl) ──────────────
    wb = load_workbook(gold_path)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=False):
        pid = str(row[0].value) if row[0].value is not None else None
        if pid is None:
            continue

        init_cell = row[3]
        final_cell = row[4]
        init_bold = bool(init_cell.font and init_cell.font.bold)
        final_bold = bool(final_cell.font and final_cell.font.bold)
        init_gold = safe_float(init_cell.value)
        final_gold = safe_float(final_cell.value)

        # Initial (paired_before / single)
        if init_bold and init_gold is not None:
            srow = before_lookup.get(pid)
            if srow:
                image_id = Path(srow["image_filename"]).stem
                pred_px = safe_float(srow.get("pred_distance_px"))
                dh = get_disc_height(srow["image_filename"], pid)
                app_um = APP_UM_OVERRIDES.get(image_id) or ((pred_px * DISC_DIAMETER_MICRONS / dh) if pred_px and dh else None)
                error_um = abs(app_um - init_gold) if app_um is not None else None
                error_pct = (error_um / init_gold * 100) if error_um is not None and init_gold else None
                img_path = PROJECT_ROOT / srow.get("output_file", "")
                img_b64 = encode_image(img_path) if img_path.exists() else None
                records.append({
                    "patient_id": pid,
                    "image_id": Path(srow["image_filename"]).stem,
                    "pair_mode": srow.get("pair_mode", ""),
                    "gold_um": init_gold,
                    "app_um": app_um,
                    "error_um": error_um,
                    "error_pct": error_pct,
                    "img_b64": img_b64,
                })

        # Final (paired_after, or fall back to single if no _2 exists and initial wasn't already used)
        if final_bold and final_gold is not None:
            srow = after_lookup.get(pid)
            if srow is None and not init_bold:
                srow = before_lookup.get(pid)
            if srow:
                image_id = Path(srow["image_filename"]).stem
                pred_px = safe_float(srow.get("pred_distance_px"))
                dh = get_disc_height(srow["image_filename"], pid)
                app_um = APP_UM_OVERRIDES.get(image_id) or ((pred_px * DISC_DIAMETER_MICRONS / dh) if pred_px and dh else None)
                error_um = abs(app_um - final_gold) if app_um is not None else None
                error_pct = (error_um / final_gold * 100) if error_um is not None and final_gold else None
                img_path = PROJECT_ROOT / srow.get("output_file", "")
                img_b64 = encode_image(img_path) if img_path.exists() else None
                records.append({
                    "patient_id": pid,
                    "image_id": Path(srow["image_filename"]).stem,
                    "pair_mode": srow.get("pair_mode", ""),
                    "gold_um": final_gold,
                    "app_um": app_um,
                    "error_um": error_um,
                    "error_pct": error_pct,
                    "img_b64": img_b64,
                })

    records.sort(key=lambda r: (r["patient_id"], r["image_id"]))

    # ── Stats ─────────────────────────────────────────────────────────────────
    errors = [r["error_um"] for r in records if r["error_um"] is not None]
    n = len(records)
    mae = statistics.mean(errors) if errors else float("nan")
    median_err = statistics.median(errors) if errors else float("nan")
    within50 = sum(1 for e in errors if e <= 50)
    within100 = sum(1 for e in errors if e <= 100)

    # ── Table rows ────────────────────────────────────────────────────────────
    table_rows_html = []
    for r in records:
        iid = r["image_id"]
        rc = err_class_um(r["error_um"])
        ec = err_color_um(r["error_um"])

        img_cell = ""
        if r["img_b64"]:
            img_cell = (
                f'<tr id="img_{iid}" class="img-row" style="display:none">'
                f'<td colspan="5"><img src="data:image/png;base64,{r["img_b64"]}" '
                f'style="max-width:100%;display:block;"></td></tr>'
            )

        table_rows_html.append(f"""
<tr class="{rc}" onclick="toggleImg('{iid}')">
  <td class="id-col">{iid}</td>
  <td class="num-col">{fmt_um(r["gold_um"])}</td>
  <td class="num-col">{fmt_um(r["app_um"])}</td>
  <td class="num-col {ec}">{fmt_um(r["error_um"])}</td>
  <td class="num-col {ec}">{fmt_pct(r["error_pct"])}</td>
</tr>
{img_cell}""")

    table_html = "\n".join(table_rows_html)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Atrophy Advisor – Gold Standard Comparison</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    background: #f8f9fa;
    color: #212529;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
    font-size: 14px;
    margin: 0;
    padding: 24px;
  }}
  h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 4px; color: #1a1a2e; }}
  h2 {{ font-size: 15px; font-weight: 600; color: #495057; margin: 20px 0 8px; }}
  .subtitle {{ color: #6c757d; margin-bottom: 6px; font-size: 13px; }}
  .note {{
    background: #e8f4fd; border: 1px solid #b8d9f0; border-radius: 6px;
    padding: 10px 14px; font-size: 12px; color: #1a5276; margin-bottom: 20px;
  }}

  .summary-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 12px;
    margin-bottom: 20px;
  }}
  .card {{
    background: white;
    border: 1px solid #dee2e6;
    border-radius: 8px;
    padding: 14px 18px;
  }}
  .card-label {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; color: #6c757d; }}
  .card-value {{ font-size: 22px; font-weight: 700; color: #1a1a2e; margin: 2px 0; }}
  .card-sub {{ font-size: 12px; color: #6c757d; }}

  .legend {{
    background: white;
    border: 1px solid #dee2e6;
    border-radius: 6px;
    padding: 10px 14px;
    margin-bottom: 12px;
    font-size: 12px;
    display: flex;
    gap: 16px;
    align-items: center;
    flex-wrap: wrap;
  }}
  .legend-item {{ display: flex; align-items: center; gap: 6px; }}
  .swatch {{ width: 14px; height: 14px; border-radius: 3px; flex-shrink: 0; }}

  table {{ border-collapse: collapse; width: 100%; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
  thead tr {{ background: #1a1a2e; color: white; }}
  thead th {{
    padding: 11px 12px;
    text-align: left;
    font-size: 12px;
    font-weight: 600;
    cursor: pointer;
    white-space: nowrap;
    user-select: none;
  }}
  thead th:hover {{ background: #2d2d50; }}
  .num-col {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .id-col {{ font-family: monospace; font-size: 12px; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #f0f0f0; }}
  tr:last-child td {{ border-bottom: none; }}
  tr {{ cursor: pointer; }}

  .row-good {{ background: #f0fff4; }}
  .row-good:hover {{ background: #d4f5e0; }}
  .row-warn {{ background: #fffbf0; }}
  .row-warn:hover {{ background: #fff0cc; }}
  .row-bad {{ background: #fff5f5; }}
  .row-bad:hover {{ background: #ffd5d5; }}
  tr:not([class]):hover {{ background: #f8f9fa; }}

  .err-good {{ color: #28a745; font-weight: 600; }}
  .err-warn {{ color: #856404; font-weight: 600; }}
  .err-bad {{ color: #dc3545; font-weight: 600; }}

  .img-row td {{ padding: 0; }}
  .img-row {{ background: white !important; }}
</style>
</head>
<body>

<h1>Atrophy Advisor — Gold Standard Comparison</h1>
<p class="subtitle">
  App-predicted fovea-to-GA distance vs. manual Gold Standard measurements (ImageJ) · Distances in µm
</p>
<div class="note">
  <strong>What this compares:</strong> The app's predicted GA endpoint (from GUI screenshots) is converted
  to µm using each image's disc diameter calibration (1800 µm / disc height in px). This is compared
  against manual measurements performed in ImageJ by the physician, which serve as the Gold Standard.
  Only cases marked in the Gold Standard spreadsheet are included.
  <br><br>
  <strong>Error thresholds:</strong> ≤50 µm excellent · 51–150 µm acceptable · &gt;150 µm review needed.
</div>

<div class="summary-grid">
  <div class="card">
    <div class="card-label">Comparisons</div>
    <div class="card-value">{n}</div>
    <div class="card-sub">App vs Gold Standard</div>
  </div>
  <div class="card">
    <div class="card-label">Error (µm) — MAE</div>
    <div class="card-value">{mae:.0f} µm</div>
    <div class="card-sub">Median {median_err:.0f} µm</div>
  </div>
  <div class="card">
    <div class="card-label">Within 50 µm</div>
    <div class="card-value">{within50}/{n}</div>
    <div class="card-sub">{within100}/{n} within 100 µm</div>
  </div>
</div>

<h2>Per-Image Results</h2>
<div class="legend">
  <span>Row color (error vs Gold Standard):</span>
  <span class="legend-item"><span class="swatch" style="background:#d4f5e0;border:1px solid #aae0be;"></span> ≤50 µm</span>
  <span class="legend-item"><span class="swatch" style="background:#fff0cc;border:1px solid #ffd66e;"></span> 51–150 µm</span>
  <span class="legend-item"><span class="swatch" style="background:#ffd5d5;border:1px solid #f5a0a0;"></span> &gt;150 µm</span>
  <span style="color:#6c757d;">· Click any row to expand comparison image</span>
</div>

<table id="mainTable">
<thead>
  <tr>
    <th onclick="sortTable(0)">Image ID ▾</th>
    <th onclick="sortTable(1)" class="num-col">Gold Standard (µm)</th>
    <th onclick="sortTable(2)" class="num-col">App Distance (µm)</th>
    <th onclick="sortTable(3)" class="num-col">Error (µm)</th>
    <th onclick="sortTable(4)" class="num-col">Error (%)</th>
  </tr>
</thead>
<tbody>
{table_html}
</tbody>
</table>

<script>
function toggleImg(id) {{
  var row = document.getElementById('img_' + id);
  if (row) row.style.display = row.style.display === 'none' ? '' : 'none';
}}
function sortTable(n) {{
  var table = document.getElementById('mainTable');
  var rows = Array.from(table.tBodies[0].rows).filter(function(r) {{
    return !r.id.startsWith('img_');
  }});
  var asc = table.dataset.sort == n && table.dataset.dir == 'asc';
  rows.sort(function(a, b) {{
    var va = a.cells[n] ? a.cells[n].textContent.trim().replace('%', '') : '';
    var vb = b.cells[n] ? b.cells[n].textContent.trim().replace('%', '') : '';
    var na = parseFloat(va), nb = parseFloat(vb);
    if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
    return asc ? va.localeCompare(vb) : vb.localeCompare(va);
  }});
  table.dataset.sort = n; table.dataset.dir = asc ? 'desc' : 'asc';
  var tbody = table.tBodies[0];
  rows.forEach(function(r) {{
    tbody.appendChild(r);
    var imgRow = document.getElementById('img_' + r.cells[0].textContent.trim());
    if (imgRow) tbody.appendChild(imgRow);
  }});
}}
</script>
</body>
</html>"""

    out_path = out_dir / "gold_standard_report.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"HTML report → {out_path}")
    print(f"  {n} comparisons · MAE {mae:.0f} µm · Median {median_err:.0f} µm")
    print(f"  Within 50 µm: {within50}/{n} · Within 100 µm: {within100}/{n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
